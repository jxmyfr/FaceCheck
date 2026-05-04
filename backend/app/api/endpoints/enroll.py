import io
import cv2
import numpy as np
import openpyxl
import zipfile
import tempfile
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Form, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session

from app.models.database import get_db, Student, StudentFaceEmbedding
from app.models.user import User
from app.services.face_proc import FaceProcessor
from app.core.dependencies import require_teacher_or_admin, require_admin

router = APIRouter()
face_processor = FaceProcessor()

FACES_DIR = Path(__file__).resolve().parents[3] / "storage" / "faces"
FACES_DIR.mkdir(parents=True, exist_ok=True)


@router.post("/register", status_code=status.HTTP_201_CREATED)
async def register_student(
    student_id:  str = Form(...),
    title:       str = Form(default=""),
    first_name:  str = Form(...),
    last_name:   str = Form(...),
    grade_level: str = Form(default=""),
    room_number: str = Form(default=""),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    if db.query(Student).filter(Student.student_id == student_id).first():
        raise HTTPException(status_code=409, detail=f"รหัสนักเรียน {student_id} มีในระบบแล้ว")

    contents = await file.read()
    nparr    = np.frombuffer(contents, np.uint8)
    frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if frame is None:
        raise HTTPException(status_code=400, detail="ไม่สามารถอ่านไฟล์ภาพได้")

    try:
        embedding = face_processor.process_capture(frame)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if embedding is None:
        raise HTTPException(status_code=400, detail="ไม่พบใบหน้าในภาพ — กรุณาให้ใบหน้าอยู่กลางภาพและมองตรงเข้าหากล้อง")

    _, jpeg_buf = cv2.imencode('.jpg', frame)
    img_bytes = jpeg_buf.tobytes()
    new_student = Student(
        student_id=student_id,
        title=title or None,
        first_name=first_name,
        last_name=last_name,
        grade_level=grade_level or None,
        room_number=room_number or None,
        face_embedding=embedding.tobytes(),
        face_image=img_bytes,
    )
    db.add(new_student)
    db.flush()
    db.add(StudentFaceEmbedding(
        student_id=new_student.id,
        embedding=embedding.tobytes(),
        face_image=img_bytes,
        label="มุมตรง",
    ))
    db.commit()
    db.refresh(new_student)
    cv2.imwrite(str(FACES_DIR / f"{student_id}.jpg"), frame)

    return {"status": "success", "message": f"ลงทะเบียน {first_name} {last_name} สำเร็จ", "id": new_student.id}


@router.put("/update-face/{student_id}")
async def update_face(
    student_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบรหัสนักเรียนในระบบ")

    contents  = await file.read()
    nparr     = np.frombuffer(contents, np.uint8)
    frame     = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if frame is None:
        raise HTTPException(status_code=400, detail="ไม่สามารถอ่านไฟล์ภาพได้")

    try:
        embedding = face_processor.process_capture(frame)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if embedding is None:
        raise HTTPException(status_code=400, detail="ไม่พบใบหน้าในภาพ — กรุณาให้ใบหน้าอยู่กลางภาพและมองตรงเข้าหากล้อง")

    _, jpeg_buf = cv2.imencode('.jpg', frame)
    student.face_embedding = embedding.tobytes()
    student.face_image = jpeg_buf.tobytes()
    db.commit()
    cv2.imwrite(str(FACES_DIR / f"{student_id}.jpg"), frame)

    return {"status": "success", "message": f"อัปเดตใบหน้าของ {student.first_name} สำเร็จ"}


@router.get("/students")
def list_students(
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    students = db.query(Student).order_by(Student.student_id).all()
    return [
        {
            "id": s.id,
            "student_id": s.student_id,
            "title": s.title,
            "first_name": s.first_name,
            "last_name": s.last_name,
            "grade_level": s.grade_level,
            "room_number": s.room_number,
            "has_face": len(s.face_embedding) > 0,
        }
        for s in students
    ]


@router.get("/students/{student_id}/face")
def get_student_face(
    student_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบรหัสนักเรียน")
    # ลองดึงจาก DB ก่อน (reliable ที่สุด)
    if student.face_image:
        return Response(content=student.face_image, media_type="image/jpeg")
    # fallback: อ่านจาก filesystem
    face_file = FACES_DIR / f"{student_id}.jpg"
    if face_file.is_file():
        return Response(content=face_file.read_bytes(), media_type="image/jpeg")
    raise HTTPException(status_code=404, detail="ไม่พบรูปใบหน้า — กรุณาอัปเดตใบหน้าใหม่")


@router.put("/students/{student_id}")
def update_student(
    student_id: str,
    title:       str = Form(default=""),
    first_name:  str = Form(...),
    last_name:   str = Form(...),
    grade_level: str = Form(default=""),
    room_number: str = Form(default=""),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบรหัสนักเรียน")
    student.title       = title or None
    student.first_name  = first_name
    student.last_name   = last_name
    student.grade_level = grade_level or None
    student.room_number = room_number or None
    db.commit()
    return {"status": "success", "message": f"อัปเดตข้อมูล {first_name} สำเร็จ"}


@router.delete("/students/{student_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_student(
    student_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบรหัสนักเรียน")

    db.delete(student)
    db.commit()

    face_file = FACES_DIR / f"{student_id}.jpg"
    if face_file.exists():
        face_file.unlink()


# ── Excel column aliases (Thai + English) ───────────────────────
_COL_MAP = {
    "รหัสนักเรียน": "student_id", "student_id": "student_id", "id": "student_id",
    "คำนำหน้า": "title",          "title": "title",
    "ชื่อ": "first_name",         "first_name": "first_name", "firstname": "first_name",
    "นามสกุล": "last_name",       "last_name": "last_name",   "lastname": "last_name",
    "ระดับชั้น": "grade_level",   "grade_level": "grade_level", "grade": "grade_level",
    "ห้อง": "room_number",        "room_number": "room_number", "room": "room_number",
}


@router.post("/register-multi", status_code=status.HTTP_201_CREATED)
async def register_student_multi(
    student_id:  str = Form(...),
    title:       str = Form(default=""),
    first_name:  str = Form(...),
    last_name:   str = Form(...),
    grade_level: str = Form(default=""),
    room_number: str = Form(default=""),
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    """ลงทะเบียนนักเรียนใหม่โดยใช้หลายรูป — average embedding"""
    if db.query(Student).filter(Student.student_id == student_id).first():
        raise HTTPException(status_code=409, detail=f"รหัสนักเรียน {student_id} มีในระบบแล้ว")

    valid_pairs = []  # [(embedding, frame, jpeg_bytes)]
    results     = []

    for f in files:
        contents = await f.read()
        nparr    = np.frombuffer(contents, np.uint8)
        frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame is None:
            results.append({"filename": f.filename, "valid": False, "reason": "อ่านไฟล์ไม่ได้"})
            continue
        try:
            embedding = face_processor.process_capture(frame)
            _, jpeg_buf = cv2.imencode('.jpg', frame)
            valid_pairs.append((embedding, frame, jpeg_buf.tobytes()))
            results.append({"filename": f.filename, "valid": True})
        except ValueError as e:
            results.append({"filename": f.filename, "valid": False, "reason": str(e)})

    if not valid_pairs:
        raise HTTPException(status_code=400, detail="ไม่มีรูปที่ผ่านการตรวจสอบ กรุณาส่งรูปใบหน้าที่ชัดเจนอย่างน้อย 1 รูป")

    valid_embeddings = [p[0] for p in valid_pairs]
    avg_emb = np.mean(valid_embeddings, axis=0)
    avg_emb = avg_emb / np.linalg.norm(avg_emb)

    best_frame, best_jpeg = valid_pairs[0][1], valid_pairs[0][2]

    new_student = Student(
        student_id=student_id,
        title=title or None,
        first_name=first_name,
        last_name=last_name,
        grade_level=grade_level or None,
        room_number=room_number or None,
        face_embedding=avg_emb.tobytes(),
        face_image=best_jpeg,
    )
    db.add(new_student)
    db.flush()  # get new_student.id before adding embeddings

    for i, (emb, _, jpeg_bytes) in enumerate(valid_pairs):
        db.add(StudentFaceEmbedding(
            student_id=new_student.id,
            embedding=emb.tobytes(),
            face_image=jpeg_bytes,
            label=f"รูปที่ {i + 1}",
        ))

    db.commit()
    db.refresh(new_student)
    cv2.imwrite(str(FACES_DIR / f"{student_id}.jpg"), best_frame)

    return {
        "status":  "success",
        "message": f"ลงทะเบียน {first_name} {last_name} สำเร็จ (ใช้ {len(valid_pairs)}/{len(files)} รูป)",
        "id":      new_student.id,
        "results": results,
    }


@router.post("/validate-photo")
async def validate_photo(
    file: UploadFile = File(...),
    _: User = Depends(require_teacher_or_admin),
):
    """ตรวจสอบคุณภาพรูปใบหน้า — ไม่บันทึกลง DB"""
    contents = await file.read()
    nparr    = np.frombuffer(contents, np.uint8)
    frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if frame is None:
        return {"valid": False, "reason": "ไม่สามารถอ่านไฟล์ภาพได้"}
    try:
        face_processor.process_capture(frame)
        return {"valid": True, "reason": "ผ่านการตรวจสอบ"}
    except ValueError as e:
        return {"valid": False, "reason": str(e)}


@router.post("/check-angle")
async def check_angle(
    file: UploadFile = File(...),
    expected: str = "front",
    _: User = Depends(require_teacher_or_admin),
):
    """ตรวจสอบมุมใบหน้าจาก yaw angle ของ InsightFace — ไม่บันทึกลง DB"""
    contents = await file.read()
    nparr    = np.frombuffer(contents, np.uint8)
    frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if frame is None:
        return {"valid": False, "detected": False, "message": "ไม่สามารถอ่านไฟล์ภาพได้"}

    faces = face_processor.app.get(frame)
    if not faces:
        return {"valid": False, "detected": False, "message": "ไม่พบใบหน้าในภาพ — ให้ใบหน้าอยู่กลางกรอบ"}

    face = max(faces, key=lambda f: (f.bbox[2] - f.bbox[0]) * (f.bbox[3] - f.bbox[1]))
    yaw  = float(face.pose[1])

    if expected == "front":
        valid = abs(yaw) <= 15
        if valid:        message = "มุมตรง ✓"
        elif yaw > 0:    message = "หันซ้ายมากไป — มองตรงเข้าหากล้อง"
        else:            message = "หันขวามากไป — มองตรงเข้าหากล้อง"
    elif expected == "left":
        valid = -50 <= yaw <= -15
        if yaw > -15:    message = "หันซ้ายให้มากกว่านี้"
        elif yaw < -50:  message = "หันซ้ายมากไป"
        else:            message = "มุมซ้าย ✓"
    elif expected == "right":
        valid = 15 <= yaw <= 50
        if yaw < 15:     message = "หันขวาให้มากกว่านี้"
        elif yaw > 50:   message = "หันขวามากไป"
        else:            message = "มุมขวา ✓"
    else:
        valid, message = True, ""

    return {"valid": valid, "detected": True, "yaw": round(yaw, 1), "message": message}


@router.put("/update-face-multi/{student_id}")
async def update_face_multi(
    student_id: str,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    """อัปเดตใบหน้าโดยใช้หลายรูป — average embedding เพื่อความแม่นยำสูงขึ้น"""
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบรหัสนักเรียนในระบบ")

    valid_embeddings = []
    best_frame       = None
    results          = []

    for f in files:
        contents = await f.read()
        nparr    = np.frombuffer(contents, np.uint8)
        frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame is None:
            results.append({"filename": f.filename, "valid": False, "reason": "อ่านไฟล์ไม่ได้"})
            continue
        try:
            embedding = face_processor.process_capture(frame)
            valid_embeddings.append(embedding)
            if best_frame is None:
                best_frame = frame
            results.append({"filename": f.filename, "valid": True, "reason": "ผ่าน"})
        except ValueError as e:
            results.append({"filename": f.filename, "valid": False, "reason": str(e)})

    if not valid_embeddings:
        raise HTTPException(status_code=400, detail="ไม่มีรูปที่ผ่านการตรวจสอบ กรุณาส่งรูปใบหน้าที่ชัดเจนอย่างน้อย 1 รูป")

    # Average and re-normalise embeddings
    avg_emb = np.mean(valid_embeddings, axis=0)
    avg_emb = avg_emb / np.linalg.norm(avg_emb)

    _, jpeg_buf = cv2.imencode('.jpg', best_frame)
    student.face_embedding = avg_emb.tobytes()
    student.face_image     = jpeg_buf.tobytes()
    db.commit()
    cv2.imwrite(str(FACES_DIR / f"{student_id}.jpg"), best_frame)

    return {
        "status":  "success",
        "message": f"อัปเดตใบหน้าสำเร็จ ใช้ {len(valid_embeddings)}/{len(files)} รูป",
        "results": results,
    }


MAX_FACE_SLOTS = 50


@router.get("/students/{student_id}/embeddings")
def list_embeddings(
    student_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบนักเรียน")
    return [
        {
            "id":         e.id,
            "label":      e.label,
            "source":     "scan" if (e.label or "").startswith("สแกน") else "manual",
            "created_at": e.created_at.strftime("%d/%m/%Y %H:%M") if e.created_at else None,
            "has_image":  bool(e.face_image),
        }
        for e in student.face_embeddings
    ]


@router.get("/students/{student_id}/embeddings/{emb_id}/image")
def get_embedding_image(
    student_id: str,
    emb_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบนักเรียน")
    emb = db.query(StudentFaceEmbedding).filter(
        StudentFaceEmbedding.id == emb_id,
        StudentFaceEmbedding.student_id == student.id,
    ).first()
    if not emb or not emb.face_image:
        raise HTTPException(status_code=404, detail="ไม่พบรูป")
    return Response(content=emb.face_image, media_type="image/jpeg")


@router.post("/students/{student_id}/embeddings", status_code=201)
async def add_embedding(
    student_id: str,
    label: str = Form(default=""),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบนักเรียน")

    slot_count = len(student.face_embeddings)
    if slot_count >= MAX_FACE_SLOTS:
        raise HTTPException(status_code=400, detail=f"มีใบหน้าครบ {MAX_FACE_SLOTS} มุมแล้ว ลบมุมเก่าก่อนเพิ่มใหม่")

    contents = await file.read()
    nparr    = np.frombuffer(contents, np.uint8)
    frame    = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if frame is None:
        raise HTTPException(status_code=400, detail="อ่านไฟล์ภาพไม่ได้")

    try:
        embedding = face_processor.process_capture(frame)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if embedding is None:
        raise HTTPException(status_code=400, detail="ไม่พบใบหน้าในภาพ")

    _, jpeg_buf = cv2.imencode('.jpg', frame)
    auto_label = label.strip() or f"รูปที่ {slot_count + 1}"

    new_emb = StudentFaceEmbedding(
        student_id=student.id,
        embedding=embedding.tobytes(),
        face_image=jpeg_buf.tobytes(),
        label=auto_label,
    )
    db.add(new_emb)

    # Update primary embedding on Student to the latest one
    student.face_embedding = embedding.tobytes()
    student.face_image     = jpeg_buf.tobytes()
    db.commit()
    db.refresh(new_emb)

    return {
        "id":    new_emb.id,
        "label": new_emb.label,
        "slots": slot_count + 1,
        "max":   MAX_FACE_SLOTS,
    }


@router.post("/students/{student_id}/embeddings/bulk", status_code=201)
async def add_embeddings_bulk(
    student_id: str,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    """เพิ่มหลายมุมใบหน้าพร้อมกัน — แต่ละไฟล์ = 1 slot"""
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบนักเรียน")

    current_slots = len(student.face_embeddings)
    remaining = MAX_FACE_SLOTS - current_slots
    if remaining <= 0:
        raise HTTPException(status_code=400, detail=f"มีใบหน้าครบ {MAX_FACE_SLOTS} มุมแล้ว")

    added = 0
    results = []

    for f in files[:remaining]:
        contents = await f.read()
        nparr = np.frombuffer(contents, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if frame is None:
            results.append({"filename": f.filename, "ok": False, "reason": "อ่านไฟล์ไม่ได้"})
            continue
        try:
            embedding = face_processor.process_capture(frame)
        except ValueError as e:
            results.append({"filename": f.filename, "ok": False, "reason": str(e)})
            continue
        if embedding is None:
            results.append({"filename": f.filename, "ok": False, "reason": "ไม่พบใบหน้า"})
            continue

        _, jpeg_buf = cv2.imencode('.jpg', frame)
        auto_label = f"รูปที่ {current_slots + added + 1}"
        db.add(StudentFaceEmbedding(
            student_id=student.id,
            embedding=embedding.tobytes(),
            face_image=jpeg_buf.tobytes(),
            label=auto_label,
        ))
        student.face_embedding = embedding.tobytes()
        student.face_image = jpeg_buf.tobytes()
        added += 1
        results.append({"filename": f.filename, "ok": True, "label": auto_label})

    if added == 0:
        raise HTTPException(status_code=400, detail="ไม่มีรูปที่ผ่านการตรวจสอบ — ตรวจสอบว่าภาพมีใบหน้าที่ชัดเจน")

    db.commit()
    return {
        "added": added,
        "total_slots": current_slots + added,
        "max": MAX_FACE_SLOTS,
        "results": results,
    }


@router.delete("/students/{student_id}/embeddings/{emb_id}", status_code=204)
def delete_embedding(
    student_id: str,
    emb_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    student = db.query(Student).filter(Student.student_id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="ไม่พบนักเรียน")
    if len(student.face_embeddings) <= 1:
        raise HTTPException(status_code=400, detail="ต้องมีใบหน้าอย่างน้อย 1 มุม — ไม่สามารถลบมุมสุดท้ายได้")
    emb = db.query(StudentFaceEmbedding).filter(
        StudentFaceEmbedding.id == emb_id,
        StudentFaceEmbedding.student_id == student.id,
    ).first()
    if not emb:
        raise HTTPException(status_code=404, detail="ไม่พบ embedding")
    db.delete(emb)
    # Update primary to first remaining
    db.flush()
    remaining = db.query(StudentFaceEmbedding).filter(
        StudentFaceEmbedding.student_id == student.id
    ).order_by(StudentFaceEmbedding.id).first()
    if remaining:
        student.face_embedding = remaining.embedding
        student.face_image     = remaining.face_image
    db.commit()


@router.get("/template")
def download_template(_: User = Depends(require_teacher_or_admin)):
    """ดาวน์โหลด Excel template สำหรับนำเข้านักเรียน"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "นักเรียน"
    headers = ["รหัสนักเรียน", "คำนำหน้า", "ชื่อ", "นามสกุล", "ระดับชั้น", "ห้อง"]
    ws.append(headers)
    # example rows
    ws.append(["6408052201", "นาย", "สมชาย", "ใจดี", "ม.5", "1"])
    ws.append(["6408052202", "นางสาว", "สมหญิง", "รักเรียน", "ม.5", "1"])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_template.xlsx"},
    )


@router.get("/students/export")
def export_students(
    grade_level: str = None,
    room_number: str = None,
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    """Export รายชื่อนักเรียนเป็น Excel"""
    q = db.query(Student).order_by(Student.grade_level, Student.room_number, Student.student_id)
    if grade_level:
        q = q.filter(Student.grade_level == grade_level)
    if room_number:
        q = q.filter(Student.room_number == room_number)
    students = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายชื่อนักเรียน"

    headers = ["รหัสนักเรียน", "คำนำหน้า", "ชื่อ", "นามสกุล", "ระดับชั้น", "ห้อง", "มีใบหน้า"]
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, s in enumerate(students, 2):
        ws.cell(row=row_num, column=1, value=s.student_id)
        ws.cell(row=row_num, column=2, value=s.title or "")
        ws.cell(row=row_num, column=3, value=s.first_name)
        ws.cell(row=row_num, column=4, value=s.last_name)
        ws.cell(row=row_num, column=5, value=s.grade_level or "")
        ws.cell(row=row_num, column=6, value=s.room_number or "")
        ws.cell(row=row_num, column=7, value="มี" if len(s.face_embedding) > 0 else "ไม่มี")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"},
    )


@router.post("/import-excel", status_code=201)
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="รองรับเฉพาะไฟล์ .xlsx เท่านั้น")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="ไม่สามารถอ่านไฟล์ Excel ได้")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="ไฟล์ว่างเปล่า")

    # Map header row
    raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_index: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        canonical = _COL_MAP.get(h) or _COL_MAP.get(str(rows[0][i]).strip() if rows[0][i] else "")
        if canonical and canonical not in col_index:
            col_index[canonical] = i

    required = {"student_id", "first_name", "last_name"}
    missing = required - col_index.keys()
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"ไม่พบคอลัมน์ที่จำเป็น: {', '.join(missing)} — ดู template สำหรับรูปแบบที่ถูกต้อง",
        )

    def cell(row, key):
        idx = col_index.get(key)
        return str(row[idx]).strip() if idx is not None and row[idx] is not None else ""

    created, duplicates, errors = [], [], []

    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue  # skip blank rows

        sid  = cell(row, "student_id")
        fn   = cell(row, "first_name")
        ln   = cell(row, "last_name")
        grade = cell(row, "grade_level")
        room  = cell(row, "room_number")

        if not sid or not fn or not ln:
            errors.append({"row": row_num, "reason": "ข้อมูลไม่ครบ (รหัส/ชื่อ/นามสกุล)"})
            continue

        if db.query(Student).filter(Student.student_id == sid).first():
            duplicates.append({"student_id": sid, "name": f"{fn} {ln}"})
            continue

        db.add(Student(
            student_id=sid,
            first_name=fn,
            last_name=ln,
            grade_level=grade or None,
            room_number=room or None,
            face_embedding=b"",
        ))
        created.append({"student_id": sid, "name": f"{fn} {ln}"})

    db.commit()
    return {
        "created":    len(created),
        "duplicates": len(duplicates),
        "errors":     len(errors),
        "created_list":    created,
        "duplicate_list":  duplicates,
        "error_list":      errors,
    }


def _import_students_from_ws(ws, db):
    """แยก logic import Excel ให้ใช้ซ้ำได้"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], [], []

    raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_index: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        canonical = _COL_MAP.get(h) or _COL_MAP.get(str(rows[0][i]).strip() if rows[0][i] else "")
        if canonical and canonical not in col_index:
            col_index[canonical] = i

    required = {"student_id", "first_name", "last_name"}
    if required - col_index.keys():
        return [], [], [], []

    def cell(row, key):
        idx = col_index.get(key)
        return str(row[idx]).strip() if idx is not None and row[idx] is not None else ""

    created, duplicates, errors, sid_list = [], [], [], []

    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        sid   = cell(row, "student_id")
        fn    = cell(row, "first_name")
        ln    = cell(row, "last_name")
        grade = cell(row, "grade_level")
        room  = cell(row, "room_number")

        if not sid or not fn or not ln:
            errors.append({"row": row_num, "reason": "ข้อมูลไม่ครบ (รหัส/ชื่อ/นามสกุล)"})
            continue

        if db.query(Student).filter(Student.student_id == sid).first():
            duplicates.append({"student_id": sid, "name": f"{fn} {ln}"})
            continue

        db.add(Student(
            student_id=sid,
            first_name=fn,
            last_name=ln,
            grade_level=grade or None,
            room_number=room or None,
            face_embedding=b"",
        ))
        created.append({"student_id": sid, "name": f"{fn} {ln}"})
        sid_list.append(sid)

    db.commit()
    return created, duplicates, errors, sid_list


@router.post("/import-zip", status_code=201)
async def import_zip(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_teacher_or_admin),
):
    if not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="รองรับเฉพาะไฟล์ .zip เท่านั้น")

    contents = await file.read()
    try:
        zf = zipfile.ZipFile(io.BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=400, detail="ไม่สามารถเปิดไฟล์ ZIP ได้")

    names = zf.namelist()

    # หา Excel file ในซิป
    excel_names = [n for n in names if n.lower().endswith((".xlsx", ".xls")) and not n.startswith("__")]
    if not excel_names:
        raise HTTPException(status_code=400, detail="ไม่พบไฟล์ Excel (.xlsx) ใน ZIP")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(excel_names[0])), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="ไม่สามารถอ่านไฟล์ Excel ใน ZIP ได้")

    created, duplicates, errors, sid_list = _import_students_from_ws(wb.active, db)

    # หารูปภาพใน ZIP — ชื่อไฟล์ต้องเป็น {student_id}.jpg/png/jpeg
    image_exts = {".jpg", ".jpeg", ".png"}
    image_files = {
        Path(n).stem: n for n in names
        if Path(n).suffix.lower() in image_exts and not n.startswith("__")
    }

    faces_ok, faces_fail = [], []

    for sid in set(sid_list) | set(image_files.keys()):
        if sid not in image_files:
            continue
        img_bytes = zf.read(image_files[sid])
        arr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if frame is None:
            faces_fail.append({"student_id": sid, "reason": "อ่านรูปไม่ได้"})
            continue

        try:
            embedding = face_processor.extract_embedding(frame)
        except Exception:
            faces_fail.append({"student_id": sid, "reason": "ตรวจจับใบหน้าไม่พบ"})
            continue

        student = db.query(Student).filter(Student.student_id == sid).first()
        if not student:
            faces_fail.append({"student_id": sid, "reason": "ไม่พบนักเรียนในระบบ"})
            continue

        _, jpeg_buf = cv2.imencode(".jpg", frame)
        student.face_embedding = embedding.tobytes()
        student.face_image = jpeg_buf.tobytes()
        faces_ok.append({"student_id": sid})

    db.commit()

    return {
        "created":     len(created),
        "duplicates":  len(duplicates),
        "errors":      len(errors),
        "faces_ok":    len(faces_ok),
        "faces_fail":  len(faces_fail),
        "created_list":   created,
        "duplicate_list": duplicates,
        "error_list":     errors,
        "faces_fail_list": faces_fail,
    }