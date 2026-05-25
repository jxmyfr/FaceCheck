import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import date

from app.models.database import get_db, AttendanceLog, Student, Subject
from app.models.user import User, TeacherSubject
from app.core.dependencies import require_teacher_or_admin

router = APIRouter()


@router.get("/export")
def export_attendance(
    date_from:   Optional[str] = Query(default=None, description="YYYY-MM-DD"),
    date_to:     Optional[str] = Query(default=None, description="YYYY-MM-DD"),
    subject_id:  Optional[int] = Query(default=None),
    grade_level: Optional[str] = Query(default=None),
    room_number: Optional[str] = Query(default=None),
    student_id:  Optional[str] = Query(default=None, description="รหัสนักเรียน"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_teacher_or_admin),
):
    q = (
        db.query(AttendanceLog, Student, Subject)
        .join(Student, Student.id == AttendanceLog.student_id)
        .join(Subject, Subject.id == AttendanceLog.subject_id)
    )
    if current_user.role == "teacher":
        q = q.join(TeacherSubject, TeacherSubject.subject_id == Subject.id).filter(
            TeacherSubject.teacher_id == current_user.id
        )
    if date_from:
        q = q.filter(func.date(AttendanceLog.timestamp) >= date_from)
    if date_to:
        q = q.filter(func.date(AttendanceLog.timestamp) <= date_to)
    if subject_id:
        q = q.filter(AttendanceLog.subject_id == subject_id)
    if grade_level:
        q = q.filter(Student.grade_level == grade_level)
    if room_number:
        q = q.filter(Student.room_number == room_number)
    if student_id:
        q = q.filter(Student.student_id == student_id)

    rows = q.order_by(AttendanceLog.timestamp.desc()).all()

    # ── Build Excel ──────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานการเช็คชื่อ"

    headers = ["รหัสนักเรียน", "ชื่อ-นามสกุล", "ชั้น", "ห้อง",
               "รหัสวิชา", "ชื่อวิชา", "วันที่", "เวลา", "สถานะ", "วิธีเช็คชื่อ", "เหตุผล"]
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    STATUS_LABEL = {"present": "มาเรียน", "late": "มาสาย", "absent": "ขาดเรียน", "excused": "ลา", "already_checked": "เช็คแล้ว"}
    METHOD_LABEL = {"face": "สแกนใบหน้า", "qr": "QR Code", "manual": "บันทึกมือ"}

    for log, student, subject in rows:
        name = " ".join(filter(None, [student.title, student.first_name, student.last_name]))
        ws.append([
            student.student_id,
            name,
            student.grade_level or "",
            student.room_number or "",
            subject.subject_code,
            subject.subject_name,
            log.timestamp.strftime("%Y-%m-%d"),
            log.timestamp.strftime("%H:%M"),
            STATUS_LABEL.get(log.status, log.status),
            METHOD_LABEL.get(log.check_method or "", log.check_method or ""),
            log.reason or "",
        ])

    col_widths = [14, 24, 8, 8, 14, 28, 12, 8, 10, 14, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("สรุป")
    ws2.append(["สรุปรายงาน"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append(["ช่วงวันที่", f"{date_from or '-'} ถึง {date_to or '-'}"])
    ws2.append(["จำนวนรายการทั้งหมด", len(rows)])
    ws2.append(["มาเรียน", sum(1 for l, _, __ in rows if l.status == "present")])
    ws2.append(["มาสาย",  sum(1 for l, _, __ in rows if l.status == "late")])
    ws2.append(["ขาดเรียน", sum(1 for l, _, __ in rows if l.status == "absent")])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 26

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"attendance_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
